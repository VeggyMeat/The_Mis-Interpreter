import tempfile
import subprocess
import platform
import os
from typing import Optional

from program.code_block import CodeBlock
from program.assignment import Assignment
from program.condition import Condition
from program.output import Output
from program.variable import Variable
from program.value import Value
from program.operation import Operation
from program.operator import Operator
from transpilers.transpiler import Transpiler

try:
    import openpyxl
except ModuleNotFoundError:
    raise ModuleNotFoundError("Please install openpyxl: pip install openpyxl")


class ExcelTranspiler(Transpiler):
    """Transpiles C-like intermediate representation into Excel formulas/code."""

    def __init__(self, code_block: CodeBlock):
        super().__init__(code_block)

    # --------------------------
    # Convert CodeBlock → Excel
    # --------------------------
    def _convert_expression(self, expr) -> str:
        if isinstance(expr, Value):
            return str(expr.value)
        elif isinstance(expr, Variable):
            return expr.name
        elif isinstance(expr, Operation):
            left = self._convert_expression(expr.left_operand)
            right = self._convert_expression(expr.right_operand)
            op = self._convert_operator(expr.operator)
            return f"({left}{op}{right})"
        else:
            raise ValueError(f"Unknown expression type: {expr}")

    def _convert_operator(self, op: Operator) -> str:
        mapping = {
            Operator.ADD: "+",
            Operator.SUBTRACT: "-",
            Operator.LEFT_SHIFT: "<<",
            Operator.RIGHT_SHIFT: ">>",
            Operator.EQUALS: "=",
            Operator.NOT_EQUALS: "<>",
            Operator.LESS_THAN: "<",
            Operator.GREATER_THAN: ">"
        }
        return mapping[op]

    def _convert_assignment(self, assignment: Assignment) -> str:
        return f"{assignment.variable.name} = {self._convert_expression(assignment.expression)}"

    def _convert_condition(self, condition: Condition) -> str:
        # Simple Excel IF placeholder; nested blocks will appear in next rows
        expr = self._convert_expression(condition.expression)
        return f"=IF({expr}, TRUE_VAL, FALSE_VAL)"

    def _convert_output(self, output: Output) -> str:
        return f"OUTPUT: {self._convert_expression(output.expression)}"

    def _write_block_to_sheet(self, code_block: CodeBlock, ws, row_counter: int) -> int:
        for cmd in code_block.commands:
            if isinstance(cmd, Assignment):
                ws.cell(row=row_counter, column=1, value=self._convert_assignment(cmd))
                row_counter += 1
            elif isinstance(cmd, Condition):
                ws.cell(row=row_counter, column=2, value=self._convert_condition(cmd))
                row_counter += 1
                if getattr(cmd, "true_code_block", None):
                    row_counter = self._write_block_to_sheet(cmd.true_code_block, ws, row_counter)
                if getattr(cmd, "false_code_block", None):
                    ws.cell(row=row_counter, column=2, value="ELSE")
                    row_counter += 1
                    row_counter = self._write_block_to_sheet(cmd.false_code_block, ws, row_counter)
            elif isinstance(cmd, Output):
                ws.cell(row=row_counter, column=3, value=self._convert_output(cmd))
                row_counter += 1
        return row_counter

    def _generate_excel_file(self) -> str:
        wb = openpyxl.Workbook()
        ws: openpyxl.worksheet.worksheet.Worksheet = wb.active
        ws.title = "Code"

        # Headers
        ws.cell(row=1, column=1, value="Assignments / Formulas")
        ws.cell(row=1, column=2, value="Conditionals / Statements")
        ws.cell(row=1, column=3, value="Outputs")

        # Fill in the code
        self._write_block_to_sheet(self.code_block, ws, row_counter=2)

        # Save to temporary file
        tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(tmp_file.name)
        tmp_file.close()
        return tmp_file.name

    # --------------------------
    # Implements Transpiler interface
    # --------------------------
    def run_in(self) -> None:
        """Generate Excel code from code block and open it as popup."""
        file_path = self._generate_excel_file()
        system = platform.system()
        if system == "Darwin":
            subprocess.Popen(["open", file_path])
        elif system == "Windows":
            os.startfile(file_path)
        else:
            subprocess.Popen(["xdg-open", file_path])

    def run_out(self, output: Optional[int] = None) -> None:
        """Display numeric/string output in Excel as a single cell."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Output"
        ws.cell(row=1, column=1, value=str(output) if output is not None else "None")

        tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(tmp_file.name)
        tmp_file.close()

        system = platform.system()
        if system == "Darwin":
            subprocess.Popen(["open", tmp_file.name])
        elif system == "Windows":
            os.startfile(tmp_file.name)
        else:
            subprocess.Popen(["xdg-open", tmp_file.name])

